import pandas as pd
import numpy as np
from .set import Set, Proset, Tiebreak_set
from .points import Point, Forced_winner
from .player import Player
from datetime import date, datetime
import openpyxl


class Match:
    type = -1
    input_map = {"f": "Forced Winner", "u": "Unforced Error", "w": "Winner"}
    team = {1: 0, 2: 0, 3: 1, 4: 1}
    team_map = {"w": 0, "f": 0, "u": 1}

    @classmethod
    def create(cls, message_type, *args, **kwargs):
        MESSAGE_TYPE_TO_CLASS_MAP = {
            0: Match_tie,
            1: Match_3_sets,
            2: Match_Proset,
        }

        if message_type not in MESSAGE_TYPE_TO_CLASS_MAP:
            raise ValueError("Bad message type {}".format(message_type))

        return MESSAGE_TYPE_TO_CLASS_MAP[message_type](*args, **kwargs)

    def __init__(
        self,
        players,
        date=date.today(),
        tournament="practice",
        r="None",
        adv_game=False,
    ) -> None:
        # Ensure date is a date object and in the consistent format
        if isinstance(date, str):
            try:
                self.date = datetime.strptime(date, "%Y-%m-%d").date()
            except ValueError:
                raise ValueError(
                    f"Invalid date format for {date}. Expected format: YYYY-MM-DD"
                )
        else:
            self.date = date

        self.tournament = tournament
        self.r = r
        self.players = players
        self.raw_score = []
        self.raw_input = []
        self.adv_game = adv_game
        self.sets = []  # changed order of this operation
        self.current_set = self.new_set()
        self.finished = False

    def __str__(self) -> str:
        match_type = self.__class__.__name__
        return f"A {match_type} {self.r} match in {self.tournament} on {self.date} between {self.players[0].name}/{self.players[1].name} vs {self.players[2].name}/{self.players[3].name}"

    def check_match_finished(self):
        team1_wins = sum(1 for s in self.sets if s.winner == 1)
        team2_wins = sum(1 for s in self.sets if s.winner == 2)
        
        # Check the match type
        if isinstance(self, Match_Proset):
            # For Proset, a team needs to win 1 set
            return team1_wins >= 1 or team2_wins >= 1
        else:
            # For other match types, a team needs to win 2 sets
            return team1_wins >= 2 or team2_wins >= 2

    def update(self, x, silent=False):
        if x[1] == "f":
            p = Forced_winner(x)
        else:
            p = Point(x)

        self.raw_score.append(
            (
                len(self.sets) + 1,
                self.current_set.score(),
                len(self.current_set.games) + 1,
                self.current_set.current_game.score,
                self.players[p.player - 1].name,
                Match.team[p.player],
                p.category,
                p.side,
                p.shot_type,
                p.direction,
                p,
            )
        )

        team_action = Match.team[int(x[0])]
        point_winner = (team_action + Match.team_map[x[1]]) % 2
        self.current_set.update(point_winner + 1, p)

        if self.current_set.finished:
            self.sets.append(self.current_set)
            self.current_set = self.new_set()
            if self.check_match_finished():
                self.finished = True
                return
            if not silent:  # Only print if silent mode is not enabled
                for i, se in enumerate(self.sets):
                    print(i, se)

    def new_set(self):
        return Set(adv_game=self.adv_game)

    def get_set_scores(self):
        return ", ".join([s.score() for s in self.sets])

    def play_match(self, list):
        for l in list:
            self.process(l)

    def process(self, l):
        from pypadel.input.input import input_ok

        if l[0] == "!":
            self.raw_input.append(l)
        elif l[0] == "#":
            self.current_set.update_server(int(l[1]))
            self.raw_input.append(l)
        else:
            while not input_ok(l):
                l = input(f"{l} is an invalid input. Please correct it:")
            self.raw_input.append(l)
            self.update(l)

    def get_points_in_category(self, category_name, player_name=None):
        """
        Returns the number of points in a specified category for a specific player during a match.

        Parameters:
        - category_name (str): Name of the category (e.g. 'Winner', 'Unforced Error', 'Forced Winner')
        - player_name (str, optional): Name of the player. If None, returns points in the category for the whole match.

        Returns:
        int: Number of points in the specified category for the specified player (or for the whole match if no player specified).
        """

        # If a player name is provided, count the points for that player
        if player_name:
            count = sum(
                1
                for raw in self.raw_score
                if raw[6] == category_name and raw[4] == player_name
            )
        # If no player name is provided, count the points for the entire match
        else:
            count = sum(1 for raw in self.raw_score if raw[6] == category_name)

        return count

    def total_games_played(self):
        total_games = 0

        def games_from_tiebreak(tiebreak_game):
            return len(tiebreak_game.points) / 5.5

        # Count games from finished sets
        for set_ in self.sets:
            # Regular games count
            total_games += len(set_.games)

            # Check for tiebreaks based on final set scores
            if (set_.score_t1 == 7 and set_.score_t2 == 6) or (
                set_.score_t1 == 6 and set_.score_t2 == 7
            ):
                total_games += games_from_tiebreak(set_.games[-1])
            elif (set_.score_t1 == 9 and set_.score_t2 == 8) or (
                set_.score_t1 == 8 and set_.score_t2 == 9
            ):
                total_games += games_from_tiebreak(set_.games[-1])
            # Check for supertiebreaks (represented as Tiebreak_set)
            elif isinstance(set_, Tiebreak_set):
                total_games += games_from_tiebreak(set_.current_game)

        return total_games

    @classmethod
    def from_record(cls, record):
        (
            id,
            date,
            tournament,
            r,
            player_1_name,
            player_2_name,
            sets_score,
            player_3_name,
            player_4_name,
            match_type,
            raw_input,
            cat,
            adv_game,
        ) = record

        raw_input_data = raw_input.split(",")

        # Check if adv_game is a string, if not convert it to bool, default to False if not present
        if isinstance(adv_game, str):
            adv_game = adv_game.lower() == "true"
        elif adv_game is not None:
            adv_game = bool(adv_game)
        else:
            adv_game = False

        player_1 = Player(player_1_name)
        player_2 = Player(player_2_name)
        player_3 = Player(player_3_name)
        player_4 = Player(player_4_name)

        m = cls.create(
            int(match_type),
            players=[player_1, player_2, player_3, player_4],
            date=date,
            tournament=tournament,
            r=r,
            adv_game=adv_game,
        )
        m.play_match(raw_input_data)

        return m

    def get_summary(self):
        def color(val):
            if val > 2:
                color = "red"
            elif val > 1:
                color = "orange"
            else:
                color = None
            return "background-color: %s" % color

        def color_good(val):
            if val > 2:
                color = "green"
            elif val > 1:
                color = "blue"
            else:
                color = None
            return "background-color: %s" % color

        df = pd.DataFrame(
            self.raw_score,
            columns=[
                "set",
                "set_score",
                "game",
                "game_score",
                "player",
                "team",
                "category",
                "side",
                "shot_type",
                "direction",
                "raw",
            ],
        )
        p = (
            pd.pivot_table(
                df,
                values="raw",
                index=["category", "team", "player"],
                columns=["set", "set_score"],
                aggfunc="count",
            )
            .fillna(0)
            .astype(int)
        )
        idx = pd.IndexSlice[p.index.get_level_values(level=0) == "Unforced Error", :]
        idxg = pd.IndexSlice[
            p.index.get_level_values(level=0).isin(["Winner", "Forced Winner"]), :
        ]
        p = p.style.applymap(color, subset=idx).applymap(color_good, subset=idxg)
        return p

    def set_summary(self):
        df = pd.DataFrame(
            self.raw_score,
            columns=[
                "set",
                "set_score",
                "game",
                "game_score",
                "player",
                "team",
                "category",
                "side",
                "shot_type",
                "direction",
                "raw",
            ],
        )
        p = (
            pd.pivot_table(
                df,
                values="raw",
                index=["category", "team", "player"],
                columns=["set"],
                aggfunc=["count"],
            )
            .fillna(0)
            .astype(int)
        )

        # Debugging Statement 1: Print the length of self.sets
        # print(f"Length of self.sets: {len(self.sets)}")

        for x in p.columns.get_level_values(1):
            # print(f"Current value of x: {x}")
            if x <= len(
                self.sets
            ):  # check to make sure x doesn't exceed the number of sets
                p["avg", x] = p["count", x] / len(self.sets[x - 1].games)

        rename_cols = {
            x: f"{x} ({self.sets[x-1].score()})"
            for x in p.columns.get_level_values(1)
            if x <= len(self.sets)
        }
        p = p.rename(columns=rename_cols).swaplevel(axis=1)

        tot = p.groupby(level=1, axis=1).sum()

        # Identification of NaN or Infinite Values
        # non_finite_values = tot.isna() | tot.isin([np.inf, -np.inf])
        # print(non_finite_values)

        # Handle the NaN or Infinite Values (Option B for example)
        tot.fillna(0, inplace=True)
        tot.replace([np.inf, -np.inf], 0, inplace=True)

        # Casting to integers
        tot = tot.astype(int)
        p[[("Match", col) for col in tot.columns]] = tot
        p["Match", "avg"] = p["Match", "count"] / sum([len(x.games) for x in self.sets])
        p = p.sort_index(axis=1)
        return p

    def get_det_summary(self, dir=False):
        if dir:
            cols = ["set", "direction"]
        else:
            cols = ["set"]
        df = pd.DataFrame(
            self.raw_score,
            columns=[
                "set",
                "set_score",
                "game",
                "game_score",
                "player",
                "team",
                "category",
                "side",
                "shot_type",
                "direction",
                "raw",
            ],
        )
        p = (
            pd.pivot_table(
                df,
                values="raw",
                index=["team", "player", "category", "side", "shot_type"],
                columns=cols,
                aggfunc="count",
            )
            .fillna(0)
            .astype(int)
        )

        def make_pretty(styler):
            for pl in self.players:
                idx = pd.IndexSlice[p.index.get_level_values(level=1) == pl.name, :]
                styler.background_gradient(axis=None, subset=idx)
            return styler

        p = p.style.pipe(make_pretty)
        return p

    def game_summary(self, set, game):
        print(f"set {set} , Game {game}")
        self.sets[set - 1].games[game - 1].game_summary()

    def export(self, file=None):
        if not file:
            p1 = self.players[0].name.replace(" ", "")
            p3 = self.players[3].name.replace(" ", "")
            file = f"out/{self.tournament}_{p1}_{p3}.xlsx"

        with pd.ExcelWriter(file, engine="xlsxwriter") as writer:
            self.get_summary().to_excel(writer, sheet_name="match_summary")
            self.set_summary().to_excel(writer, sheet_name="set_summary")
            self.get_det_summary().to_excel(writer, sheet_name="shots_summary")
            self.get_det_summary(dir=True).to_excel(
                writer, sheet_name="shot_dir_summary"
            )

    def export_raw(self, file, sheetname="Sheet1"):
        # Attempt to load the existing workbook
        try:
            wb = openpyxl.load_workbook(file)

        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            wb = openpyxl.Workbook()

        # Check if the 'Sheet1' worksheet exists
        if sheetname in wb.sheetnames:
            # If it exists, access it for modification
            ws = wb[sheetname]
        else:
            # If it doesn't exist, create a new worksheet
            ws = wb.active
            ws.title = sheetname

        r = ws.max_row + 1
        ws.cell(column=1, row=r, value=self.date)
        ws.cell(column=2, row=r, value=self.tournament)
        ws.cell(column=3, row=r, value=str(self.r))
        ws.cell(column=4, row=r, value=str(self.players[0].name))
        ws.cell(column=5, row=r, value=str(self.players[1].name))
        ws.cell(column=6, row=r, value=str(self.players[2].name))
        ws.cell(column=7, row=r, value=str(self.players[3].name))
        ws.cell(column=8, row=r, value=self.type)
        ws.cell(column=9, row=r, value=str(",".join(self.raw_input)))
        ws.cell(column=10, row=r, value=self.adv_game)
        wb.save(file)


class Match_tie(Match):
    type = 0

    def __init__(
        self,
        players,
        date=date.today(),
        tournament="practise",
        r="None",
        adv_game=False,
    ) -> None:
        super().__init__(players, date, tournament, r, adv_game=adv_game)

    def new_set(self):
        if len(self.sets) == 2:
            return Tiebreak_set(target=10)
        else:
            return Set(adv_game=self.adv_game)


class Match_3_sets(Match):
    type = 1

    def __init__(
        self,
        players,
        date=date.today(),
        tournament="practise",
        r="None",
        adv_game=False,
    ) -> None:
        super().__init__(players, date, tournament, r, adv_game=adv_game)

    def new_set(self):
        return Set(adv_game=self.adv_game)


class Match_Proset(Match):
    type = 2

    def __init__(
        self, players, date=date.today(), tournament="practise", r="None", adv_game=True
    ) -> None:
        super().__init__(players, date, tournament, r, adv_game=adv_game)

    def new_set(self):
        return Proset()
