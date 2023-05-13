import pandas as pd
from set import *
from points import *
from datetime import date
import openpyxl

class match:
    type = -1
    #Convention -> player 1 and 3 on the left, player 2 and 4 on the right
    input_map = {'f':'Forced Error', 'u':'Unforced Error', 'w':'Winner'}
    team = {1:0, 2:0, 3:1, 4:1}
    team_map = {'w':0,'f':0,'u':1}

    @classmethod
    def create(cls, message_type, *args,**kwargs):
        MESSAGE_TYPE_TO_CLASS_MAP = {
        0:  match_tie,
        1: match_3_sets,
        }

        if message_type not in MESSAGE_TYPE_TO_CLASS_MAP:
            raise ValueError('Bad message type {}'.format(message_type))
    
        return MESSAGE_TYPE_TO_CLASS_MAP[message_type](*args,**kwargs)
    
    def __init__(self, players, date=date.today(), tournament='practise',r='None') -> None:
        self.date = date
        self.tournament = tournament
        self.r = r
        self.players = players
        self.raw_score = []
        self.raw_input = []
        self.current_set = set()
        self.sets = []
        self.finished = False

    def __str__(self) -> str:
        return f'A {self.r} match in {self.tournament} on {self.date} between {self.players[0].name}/{self.players[1].name} vs {self.players[2].name}/{self.players[3].name}'

    def update(self, x):
        if x[1] == 'f':
            p = forced_winner(x)
        else:
            p = point(x)
        self.raw_score.append(
            (len(self.sets)+1,self.current_set.score(),
             len(self.current_set.games)+1,
             self.current_set.current_game.score(),
             self.players[p.player-1].name,
             match.team[p.player],
             p.category,
             p.side,
             p.shot_type,
             p.direction,
             p))
        team_action = match.team[int(x[0])]
        point_winner = (team_action+match.team_map[x[1]])%2
        self.current_set.update(point_winner+1, p)
        if self.current_set.finished:
            self.sets.append(self.current_set)
            self.current_set = self.new_set()
            for i,se in enumerate(self.sets):
                print(i, se)
    
    def new_set(self):
        return set()

    def play_match(self, list):
        for l in list:
            self.process(l)

    def process(self, l):
        from input import input_ok
        if l[0] == "!":
            self.raw_input.append(l)
        elif l[0] == '#':
            self.current_set.update_server(int(l[1]))
            self.raw_input.append(l)
        else:
            while not input_ok(l):
                l = input(f'{l} is an invalid input. Please correct it:')
            self.raw_input.append(l)
            self.update(l)

    def get_summary(self):
        def color(val):
            if val > 2:
                color = 'red'
            elif val > 1:
                color = 'orange'
            else:
                color = None
            return 'background-color: %s' % color
        def color_good(val):
            if val > 2:
                color = 'green'
            elif val > 1:
                color = 'blue'
            else:
                color = None
            return 'background-color: %s' % color
        df = pd.DataFrame(self.raw_score, columns=['set','set_score','game','game_score','player','team','category','side','shot_type','direction','raw'])
        p = pd.pivot_table(df, values='raw',index=['category','team','player'],columns=['set','set_score'], aggfunc='count').fillna(0).astype(int)
        idx = pd.IndexSlice[p.index.get_level_values(level=0)=='Unforced Error', :]
        idxg = pd.IndexSlice[p.index.get_level_values(level=0).isin(['Winner','Forced Winner']), :]
        p = p.style.applymap(color, subset=idx).applymap(color_good, subset=idxg)
        return p
    
    def set_summary(self):
        df = pd.DataFrame(self.raw_score, columns=['set','set_score','game','game_score','player','team','category','side','shot_type','direction','raw'])
        p = pd.pivot_table(df, values='raw',index=['category','team','player'],columns=['set'], aggfunc=['count']).fillna(0).astype(int)
        for x in p.columns.get_level_values(1):
            p['avg', x] = p['count',x]/len(self.sets[x-1].games)
        p = p.rename(columns={x:f'{x} ({self.sets[x-1].score()})' for x in p.columns.get_level_values(1)}).swaplevel(axis=1)
        tot = p.groupby(level=1, axis=1).sum().astype(int)
        p[[("Match", col) for col in tot.columns]] = tot
        p["Match", 'avg'] = p["Match", 'count']/sum([len(x.games) for x in self.sets])
        p = p.sort_index(axis=1)
        return p

    def get_det_summary(self, dir=False):
        if dir:
            cols = ['set','direction']
        else:
            cols = ['set']
        df = pd.DataFrame(self.raw_score, columns=['set','set_score','game','game_score','player','team','category','side','shot_type','direction','raw'])
        p = pd.pivot_table(df, values='raw',index=['team','player','category','side','shot_type'],columns=cols, aggfunc='count').fillna(0).astype(int)
        def make_pretty(styler):
            for pl in self.players:
                idx = pd.IndexSlice[p.index.get_level_values(level=1)==pl.name, :]
                styler.background_gradient(axis=None,subset=idx)
            return styler
        p = p.style.pipe(make_pretty)
        return p

    def game_summary(self, set, game):
        print(f'Set {set} , Game {game}')
        self.sets[set-1].games[game-1].game_summary()

    def export(self, file=None):
        if not file:
            p1 = self.players[0].name.replace(' ','')
            p3 = self.players[3].name.replace(' ','')
            file = f'out/{self.tournament}_{p1}_{p3}.xlsx'
        
        with pd.ExcelWriter(file, engine='xlsxwriter') as writer:  
            self.get_summary().to_excel(writer,sheet_name='match_summary')
            self.set_summary().to_excel(writer, sheet_name='set_summary')
            self.get_det_summary().to_excel(writer,sheet_name='shots_summary')
            self.get_det_summary(dir=True).to_excel(writer,sheet_name='shot_dir_summary') 

    def export_raw(self, file, sheetname='Sheet1'):
        
        # Attempt to load the existing workbook
        try:
            wb = openpyxl.load_workbook(file)

        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            wb = openpyxl.Workbook()

        # Check if the 'Sheet1' worksheet exists
        if sheetname in wb.sheetnames:
            # If it exists, access it for modification
            ws = wb[sheetname]
        else:
            # If it doesn't exist, create a new worksheet
            ws = wb.active
            ws.title = sheetname

        r = ws.max_row+1
        ws.cell(column=1, row=r, value=self.date)
        ws.cell(column=2, row=r, value=self.tournament)
        ws.cell(column=3, row=r, value=str(self.r))
        ws.cell(column=4, row=r, value=str(self.players[0].name))
        ws.cell(column=5, row=r, value=str(self.players[1].name))
        ws.cell(column=6, row=r, value=str(self.players[2].name))
        ws.cell(column=7, row=r, value=str(self.players[3].name))
        ws.cell(column=8, row=r, value=self.type)
        ws.cell(column=9, row=r, value=str(','.join(self.raw_input)))
        wb.save(file)
    
class match_tie(match):
    type = 0
    def __init__(self, players, date=date.today(), tournament='practise', r='None') -> None:
        super().__init__(players, date, tournament, r)
    
    def new_set(self):
        if len(self.sets) == 2:
            return tiebreak_set(target=10)
        else:
            return set()


class match_3_sets(match):
    type = 1
    def __init__(self, players, date=date.today(), tournament='practise', r='None') -> None:
        super().__init__(players, date, tournament, r)
    
    def new_set(self):
        return set()